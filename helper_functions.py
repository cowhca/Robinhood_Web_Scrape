# -*- coding: utf-8 -*-

from openpyxl.styles import Font
import openpyxl

def get_col_index(index):
    return str(openpyxl.utils.get_column_letter(index))

def set_column_width(col, width, sheet):
    sheet.column_dimensions[col].width = width

# CATEGORIES
# Creating a display for each category
def make_display(_br, _bc, _list, name, sheet, nrows, data_start_row, data_end_row):
    """
    

    Parameters
    ----------
    _br : int
        The row where the program will start writing to
    _bc : int
        The column where the program will start writing to
    _list : list
        The list of stocks tickers being displayed
    name : string
        Name to be displayed with the display
    sheet : worksheet
        Worksheet object to write to
    nrows : int
        Number of rows constituting the raw data
    data_start_row : int
        Starting row of the raw data
    data_end_row : TYPE
        Ending row of the raw data

    Returns
    -------
    None.

    """
    # Creates a display for a given type of stock, specified by the name parameter
    smaller_italic_font = Font(size=12, italic=True)
    column_title_font = Font(size = 12, name = 'Cambria')
    
    base_row = _br
    base_column = _bc
    # Putting in headers
    sheet.cell(row = base_row, column = base_column).value = name +':'
    sheet.cell(row = base_row, column = base_column + 1).value = "Shares"
    sheet.cell(row = base_row, column = base_column + 2).value = "Average Cost"
    sheet.cell(row = base_row, column = base_column + 3).value = "Price"
    sheet.cell(row = base_row, column = base_column + 4).value = "% Change"

    sheet.cell(row = base_row, column = base_column).font     = column_title_font
    sheet.cell(row = base_row, column = base_column + 1).font = column_title_font
    sheet.cell(row = base_row, column = base_column + 2).font = column_title_font
    sheet.cell(row = base_row, column = base_column + 3).font = column_title_font
    sheet.cell(row = base_row, column = base_column + 4).font = column_title_font
    if len(_list) > 0:
        t_price = 0
        t_cost = 0

        # Inputing data
        for i in range(len(_list)):
            sheet.cell(row = base_row + i + 1, column = base_column).value = _list[i]
            
            # Need to find the row of each stock
            true_row = None # true_row reperesents the row of my sheet that I keep the stock that we are looking at
            for r in range(data_start_row, data_end_row + 1):
                if sheet.cell(row = r, column = 2).value == _list[i]:
                    true_row = r
                    break
            sheet.cell(row = base_row + i + 1, column = base_column + 1).value = sheet.cell(row = true_row, column = 3).value # Shares
            
            sheet.cell(row = base_row + i + 1, column = base_column + 2).value = sheet.cell(row = true_row, column = 5).value # Average Cost
            # Now add the cost so we can present some summary statistics
            # t_cost += shares * average_cost
            t_cost += sheet.cell(row = true_row, column = 3).value * sheet.cell(row = true_row, column = 5).value
            
            sheet.cell(row = base_row + i + 1, column = base_column + 3).value = sheet.cell(row = true_row, column = 4).value # Price
            # t_price += shares * price
            t_price += sheet.cell(row = true_row, column = 3).value * sheet.cell(row = true_row, column = 4).value
            
            sheet.cell(row = base_row + i + 1, column = base_column + 4).value = sheet.cell(row = true_row, column = 8).value # Percent Change


        sheet.cell(row = base_row + len(_list) + 1, column = base_column + 2).value = t_cost
        sheet.cell(row = base_row + len(_list) + 1, column = base_column + 2).font = smaller_italic_font
        
        sheet.cell(row = base_row + len(_list) + 1, column = base_column + 3).value = t_price
        sheet.cell(row = base_row + len(_list) + 1, column = base_column + 3).font = smaller_italic_font

        sheet.cell(row = base_row + len(_list) + 1, column = base_column + 4).value = round(100*(t_price - t_cost)/t_cost, 2)
        sheet.cell(row = base_row + len(_list) + 1, column = base_column + 4).font = smaller_italic_font
    else:
        sheet.cell(row = base_row + len(_list) + 1, column = base_column + 2).value = 0
        sheet.cell(row = base_row + len(_list) + 1, column = base_column + 2).font = smaller_italic_font
        
        sheet.cell(row = base_row + len(_list) + 1, column = base_column + 3).value = 0
        sheet.cell(row = base_row + len(_list) + 1, column = base_column + 3).font = smaller_italic_font

        sheet.cell(row = base_row + len(_list) + 1, column = base_column + 4).value = 0
        sheet.cell(row = base_row + len(_list) + 1, column = base_column + 4).font = smaller_italic_font
        
def write_detailed_distribution(bc, sheet, designations, nrows, data_start_row, data_end_row):
    """
    

    Parameters
    ----------
    bc : int
        The starting column to write from
    sheet : worksheet
        The worksheet object to write to
    designations : shelve.DbfilenameShelf
        Database to designate how to categorize each stock
    nrows : int
        Number of rows in the raw data
    data_start_row : int
        Starting row of the raw data
    data_end_row : TYPE
        Ending row of the raw data

    Returns
    -------
    None.

    """
    
    title_font = Font(size=22, name = 'Cambria', bold = True)
    header_row      = 2
    start_row       = 3
    type_c          = bc
    equity_c        = bc + 1
    percent_c       = bc + 2
    ideal_c         = bc + 3
    difference_c    = bc + 4
    to_add_c        = bc + 5
    new_equity_c    = bc + 6
    new_percent_c   = bc + 7
    new_difference_c= bc + 8

    sheet.cell(row = start_row - 2, column = type_c).value = "Detailed Distribution"
    sheet.cell(row = start_row - 2, column = type_c).font = title_font

    US = 0
    lrg_b = 0
    lrg_g = 0
    lrg_v = 0
    mid_b = 0
    mid_g = 0
    mid_v = 0
    sml_b = 0
    sml_g = 0
    sml_v = 0
    inter = 0
    i_lrg_b = 0
    i_lrg_g = 0
    i_lrg_v = 0
    i_mid_b = 0
    i_mid_g = 0
    i_mid_v = 0
    i_sml_b = 0
    i_sml_g = 0
    i_sml_v = 0
    emerg = 0
    lt = 0
    it = 0
    st = 0
    bil = 0
    com = 0
    REIT = 0
    gld = 0
    misc = 0
    crypto = 0

    # Going through each stock and adding its equity to the appropriate category
    for row in range(data_start_row, data_end_row + 1):
        ticker = sheet.cell(row = row, column = 2).value
        try:
            if(designations[ticker][2] == 'TSM'):
                US += sheet.cell(row = row, column = 7).value
            elif(designations[ticker][2] == 'LCB'):
                lrg_b += sheet.cell(row = row, column = 7).value
            elif(designations[ticker][2] == 'LCV'):
                lrg_v += sheet.cell(row = row, column = 7).value
            elif(designations[ticker][2] == 'LCG'):
                lrg_g += sheet.cell(row = row, column = 7).value
            elif(designations[ticker][2] == 'MCB'):
                mid_b += sheet.cell(row = row, column = 7).value
            elif(designations[ticker][2] == 'MCV'):
                mid_v += sheet.cell(row = row, column = 7).value
            elif(designations[ticker][2] == 'MCG'):
                mid_g += sheet.cell(row = row, column = 7).value
            elif(designations[ticker][2] == 'SCB'):
                sml_b += sheet.cell(row = row, column = 7).value
            elif(designations[ticker][2] == 'SCV'):
                sml_v += sheet.cell(row = row, column = 7).value
            elif(designations[ticker][2] == 'SCG'):
                sml_g += sheet.cell(row = row, column = 7).value
            elif(designations[ticker][2] == 'TSMI'):
                inter += sheet.cell(row = row, column = 7).value
            elif(designations[ticker][2] == 'LCBI'):
                i_lrg_b += sheet.cell(row = row, column = 7).value
            elif(designations[ticker][2] == 'LCVI'):
                i_lrg_v += sheet.cell(row = row, column = 7).value
            elif(designations[ticker][2] == 'LCGI'):
                i_lrg_g += sheet.cell(row = row, column = 7).value
            elif(designations[ticker][2] == 'MCBI'):
                i_mid_b += sheet.cell(row = row, column = 7).value
            elif(designations[ticker][2] == 'MCVI'):
                i_mid_v += sheet.cell(row = row, column = 7).value
            elif(designations[ticker][2] == 'MCGI'):
                i_mid_g += sheet.cell(row = row, column = 7).value
            elif(designations[ticker][2] == 'SCBI'):
                i_sml_b += sheet.cell(row = row, column = 7).value
            elif(designations[ticker][2] == 'SCVI'):
                i_sml_v += sheet.cell(row = row, column = 7).value
            elif(designations[ticker][2] == 'SCGI'):
                i_sml_g += sheet.cell(row = row, column = 7).value
            elif(designations[ticker][2] == 'EM'):
                emerg += sheet.cell(row = row, column = 7).value
            elif(designations[ticker][2] == 'LTB'):
                lt += sheet.cell(row = row, column = 7).value
            elif(designations[ticker][2] == 'ITB'):
                it += sheet.cell(row = row, column = 7).value
            elif(designations[ticker][2] == 'STB'):
                st += sheet.cell(row = row, column = 7).value
            elif(designations[ticker][2] == 'TB'):
                bil += sheet.cell(row = row, column = 7).value
            elif(designations[ticker][2] == 'COM'):
                com += sheet.cell(row = row, column = 7).value
            elif(designations[ticker][2] == 'REIT'):
                REIT += sheet.cell(row = row, column = 7).value
            elif(designations[ticker][2] == 'GLD'):
                gld += sheet.cell(row = row, column = 7).value
            elif(designations[ticker][2] == 'CRYP'):
                crypto += sheet.cell(row = row, column = 7).value
            else:
                misc += sheet.cell(row = row, column = 7).value
        except Exception as e:
            print(e)
            pass

    sheet.cell(row = header_row, column = type_c).value = 'Type'
    sheet.cell(row = header_row, column = equity_c).value = 'Equity'
    sheet.cell(row = header_row, column = percent_c).value = '%'
  
    sheet.cell(row = start_row, column = type_c).value = 'United States:'
    sheet.cell(row = start_row + 1, column = type_c).value = 'Large Blend:'
    sheet.cell(row = start_row + 2, column = type_c).value = 'Large Value:'
    sheet.cell(row = start_row + 3, column = type_c).value = 'Large Growth:'
    sheet.cell(row = start_row + 4, column = type_c).value = 'Mid Blend:'
    sheet.cell(row = start_row + 5, column = type_c).value = 'Mid Value:'
    sheet.cell(row = start_row + 6, column = type_c).value = 'Mid Growth:'
    sheet.cell(row = start_row + 7, column = type_c).value = 'Small Blend:'
    sheet.cell(row = start_row + 8, column = type_c).value = 'Small Value:'
    sheet.cell(row = start_row + 9, column = type_c).value = 'Small Growth:'
    sheet.cell(row = start_row + 10, column = type_c).value = 'Total Inter.:'
    sheet.cell(row = start_row + 11, column = type_c).value = 'ex-US Large Blend:'
    sheet.cell(row = start_row + 12, column = type_c).value = 'ex-US Large Value:'
    sheet.cell(row = start_row + 13, column = type_c).value = 'ex-US Large Growth:'
    sheet.cell(row = start_row + 14, column = type_c).value = 'ex-US Mid Blend:'
    sheet.cell(row = start_row + 15, column = type_c).value = 'ex-US Mid Value:'
    sheet.cell(row = start_row + 16, column = type_c).value = 'ex-US Mid Growth:'
    sheet.cell(row = start_row + 17, column = type_c).value = 'ex-US Small Blend:'
    sheet.cell(row = start_row + 18, column = type_c).value = 'ex-US Small Value:'
    sheet.cell(row = start_row + 19, column = type_c).value = 'ex-US Small Growth:'
    sheet.cell(row = start_row + 20, column = type_c).value = 'Emerging:'
    sheet.cell(row = start_row + 21, column = type_c).value = 'Long Term Bonds:'
    sheet.cell(row = start_row + 22, column = type_c).value = 'Intermediate Bonds:'
    sheet.cell(row = start_row + 23, column = type_c).value = 'Short Term Bonds:'
    sheet.cell(row = start_row + 24, column = type_c).value = 'Treasury Bills:'
    sheet.cell(row = start_row + 25, column = type_c).value = 'Commodities'
    sheet.cell(row = start_row + 26, column = type_c).value = 'REITs'
    sheet.cell(row = start_row + 27, column = type_c).value = 'Gold'
    sheet.cell(row = start_row + 28, column = type_c).value = 'Crypto'
    sheet.cell(row = start_row + 29, column = type_c).value = 'Misc.'

    sheet.cell(row = start_row, column = equity_c).value = US
    sheet.cell(row = start_row + 1, column = equity_c).value = lrg_b
    sheet.cell(row = start_row + 2, column = equity_c).value = lrg_v
    sheet.cell(row = start_row + 3, column = equity_c).value = lrg_g
    sheet.cell(row = start_row + 4, column = equity_c).value = mid_b
    sheet.cell(row = start_row + 5, column = equity_c).value = mid_v
    sheet.cell(row = start_row + 6, column = equity_c).value = mid_g
    sheet.cell(row = start_row + 7, column = equity_c).value = sml_b
    sheet.cell(row = start_row + 8, column = equity_c).value = sml_v
    sheet.cell(row = start_row + 9, column = equity_c).value = sml_g
    sheet.cell(row = start_row + 10, column = equity_c).value = inter
    sheet.cell(row = start_row + 11, column = equity_c).value = i_lrg_b
    sheet.cell(row = start_row + 12, column = equity_c).value = i_lrg_v
    sheet.cell(row = start_row + 13, column = equity_c).value = i_lrg_g
    sheet.cell(row = start_row + 14, column = equity_c).value = i_mid_g
    sheet.cell(row = start_row + 15, column = equity_c).value = i_mid_v
    sheet.cell(row = start_row + 16, column = equity_c).value = i_mid_g
    sheet.cell(row = start_row + 17, column = equity_c).value = i_sml_b
    sheet.cell(row = start_row + 18, column = equity_c).value = i_sml_v
    sheet.cell(row = start_row + 19, column = equity_c).value = i_sml_g
    sheet.cell(row = start_row + 20, column = equity_c).value = emerg
    sheet.cell(row = start_row + 21, column = equity_c).value = lt
    sheet.cell(row = start_row + 22, column = equity_c).value = it
    sheet.cell(row = start_row + 23, column = equity_c).value = st
    sheet.cell(row = start_row + 24, column = equity_c).value = bil
    sheet.cell(row = start_row + 25, column = equity_c).value = com
    sheet.cell(row = start_row + 26, column = equity_c).value = REIT
    sheet.cell(row = start_row + 27, column = equity_c).value = gld
    sheet.cell(row = start_row + 28, column = equity_c).value = crypto
    sheet.cell(row = start_row + 29, column = equity_c).value = misc
    sheet.cell(row = start_row + 30, column = equity_c).value = '=SUM(' + get_col_index(equity_c) + '3:' + get_col_index(equity_c) + '32)'

    for r in range(3, 33): # Set to zero before giving specific values
        sheet.cell(row = r, column = ideal_c).value = 0
    # Making a colmun to display the distribution that I want
    sheet.cell(row = header_row, column = ideal_c).value = 'Ideal'
    
    # This is a sample distribution
    sheet.cell(row = start_row, column = ideal_c).value = 10
    sheet.cell(row = start_row + 4, column = ideal_c).value = 5
    sheet.cell(row = start_row + 5, column = ideal_c).value = 20
    sheet.cell(row = start_row + 7, column = ideal_c).value = 20
    sheet.cell(row = start_row + 8, column = ideal_c).value = 25
    sheet.cell(row = start_row + 10, column = ideal_c).value = 5
    sheet.cell(row = start_row + 20, column = ideal_c).value = 15

    
    
    # Finding each classes percentage of portfolio
    column_symbol = get_col_index(equity_c)
    for r in range(3, 33):
        sheet.cell(row = r, column = percent_c).value = '=' + '100*'+column_symbol + str(r) +'/$' + column_symbol + '$33'
    

    # Multiplying and rounding each number also by using the =ROUND... formate it can be interactive after creation
    # for r in range(3, 32):
    #     sheet.cell(row = r, column = bc + 3).value = '=ROUND('+ get_col_index(bc + 2) + str(r) + ',0)'    

    # sheet.cell(row = start_row + 29, column = bc + 3).value = '=SUM('+ get_col_index(bc+3) + '3:'+ get_col_index(bc+3) + '31)'

    # Adding a column to display difference between ideal and real percentage
    sheet.cell(row = header_row, column = difference_c).value = 'Differance'
    for r in range(3, 33):
        sheet.cell(row = r, column = difference_c).value = '=ROUND(' + get_col_index(ideal_c) + str(r) + '-' + get_col_index(percent_c) + str(r) + ', 2)'

    ''' A list of lists
    for each tuple:
        0 element: stock type
        1 element: current equity
        2 element: ideal percentage
        3 element: difference between ideal and actual
        4 element: toAdd
    '''
    differences = [['',0,0,0,0], ['',0,0,0,0], ['',0,0,0,0], ['',0,0,0,0], ['',0,0,0,0], ['',0,0,0,0], ['',0,0,0,0], ['',0,0,0,0], ['',0,0,0,0], ['',0,0,0,0], ['',0,0,0,0], ['',0,0,0,0], ['',0,0,0,0], ['',0,0,0,0], ['',0,0,0,0], ['',0,0,0,0], ['',0,0,0,0], ['',0,0,0,0], ['',0,0,0,0], ['',0,0,0,0], ['',0,0,0,0], ['',0,0,0,0], ['',0,0,0,0], ['',0,0,0,0], ['',0,0,0,0], ['',0,0,0,0], ['',0,0,0,0], ['',0,0,0,0], ['',0,0,0,0], ['',0,0,0,0]]
    
    for i in range(3, 33): # zeroing toAdd column before updating it
        sheet.cell(row = i, column = to_add_c).value = 0

    sum = 0 # variable for total equity
    # initialize list
    r = 3
    for i in differences:
        # look for that row in excel sheet
        i[0] = sheet.cell(row = r, column = type_c).value 
        i[1] = sheet.cell(row = r, column = equity_c).value 
        sum += i[1] 
        i[2] = sheet.cell(row = r, column = ideal_c).value
        r += 1


    # update differences
    def update():
        for i in differences:
            i[3] = i[2] - (100 * (i[1] + i[4]) / sum)

    update()

    toInvest = float(input("How much would you like to invest?\n"))
    increment = 0.01 # How much to add when when doing each loop of optimization
    while (toInvest > increment):
        # find largest difference (ie the stock that I should most invest in)
        maxIndex = 0
        max = 0
        for i in range(0, len(differences)):
            if differences[i][3] > max:
                max = differences[i][3]
                maxIndex = i
        differences[maxIndex][4] += increment
        toInvest -= increment
        sum += increment
        update()
    ####### This section was supposed to allow me to enter a negative number and then it would spit out #######
    ####### how much to sell but I decided that I wouldn't be selling very often so I abondoned it      #######
    # if (toInvest < 0): # If toInvest is negative that means that we are selling
    #     increment = -increment
    #     # Subtract `increment` to the stock type with smallest difference until we have no money left
    #     while (toInvest < increment):
    #         # find smallest difference (ie the stock that I should most sell)
    #         minIndex = 0
    #         minValue = 0
    #         for i in range(0, len(differences)):
    #             if differences[i][3] < minValue:
    #                 minValue = differences[i][3]
    #                 minIndex = i
    #         differences[minIndex][4] += increment
    #         toInvest -= increment
    #         sum -= increment # update total equity
    #         update()
    # else: # We are buying
    #     while (toInvest > increment):
    #         # find largest difference (ie the stock that I should most invest in)
    #         maxIndex = 0
    #         max = 0
    #         for i in range(0, len(differences)):
    #             if differences[i][3] > max:
    #                 max = differences[i][3]
    #                 maxIndex = i
    #         differences[maxIndex][4] += increment
    #         toInvest -= increment
    #         sum += increment
    #         update()
    ##############
            
    # Summing the absolute values of the differences
    sheet.cell(row = start_row + 30, column = difference_c).value = "=SUMIF(" + get_col_index(difference_c) + str(start_row) + ":" + get_col_index(difference_c) + str(start_row + 29) + "," + "\">0\") - SUMIF(" + get_col_index(difference_c) + str(start_row) + ":" + get_col_index(difference_c) + str(start_row + 28) + ",\"<0\")"

    sheet.cell(row = header_row, column = to_add_c).value = "To Add"
    r = start_row
    for i in differences:
        sheet.cell(row = r, column = to_add_c).value = i[4]
        r += 1
    # Adding a sum of our `To Add` column
    sheet.cell(row = start_row + 30, column = to_add_c).value = "=SUM(" + get_col_index(to_add_c) + str(start_row) + ":" + get_col_index(to_add_c) + str(start_row + 29) + ")"
    

    sheet.cell(row = 2, column = new_equity_c).value = 'New Equity'
    # Summing our `To Add` column with our `Equity` column to create the `New Equity` column
    for i in range(start_row, len(differences) + start_row) :
        sheet.cell(row = i, column = new_equity_c).value = "=" + get_col_index(to_add_c) + str(i) + "+" + get_col_index(equity_c) + str(i)
    # Adding a sum of our new total equity
    sheet.cell(row = start_row + 30, column = new_equity_c).value = "=SUM(" + get_col_index(new_equity_c) + str(start_row) + ":" + get_col_index(new_equity_c) + str(start_row + 29) + ")"

    sheet.cell(row = 2, column = new_percent_c).value = 'New %'
    # Populating the `New %` column
    for i in range(start_row, len(differences) + start_row) :
        sheet.cell(row = i, column = new_percent_c).value = "=round(100 *" + get_col_index(new_equity_c) + str(i) + "/" + get_col_index(new_equity_c) + str(start_row + 30) + ", 3)"
    
    sheet.cell(row = 2, column = new_difference_c).value = 'New Difference'
    # Calculating new differences
    for i in range(start_row, len(differences) + start_row):
        sheet.cell(row = i, column = new_difference_c).value = "=ROUND(" + get_col_index(ideal_c) + str(i) + "-" + get_col_index(new_percent_c) + str(i) + ", 2)"
    # Summing the absolute values of the new differences
    sheet.cell(row = start_row + 30, column = new_difference_c).value = "=SUMIF(" + get_col_index(new_difference_c) + str(start_row) + ":" + get_col_index(new_difference_c) + str(start_row + 29) + "," + "\">0\") - SUMIF(" + get_col_index(new_difference_c) + str(start_row) + ":" + get_col_index(new_difference_c) + str(start_row + 29) + ",\"<0\")"
    
    
    # Color the rows red that need the most investment
    from openpyxl.formatting.rule import ColorScaleRule
    rule = ColorScaleRule(start_type='percentile', start_value=50, start_color='ffffff',
                            end_type='percentile', end_value=100, end_color='800000')
    sheet.conditional_formatting.add('Z3:Z32', rule)
