#! /usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Tue Dec  1 09:16:08 2020

@author: conno
"""


#########################################
###### Installing/loading libraries #####
#########################################

try:
    import os
except:
    print("os library needs to be installed.\n")
    os.system("python -m pip install os")
    import os
try:
    from selenium import webdriver
    from selenium.common.exceptions import NoSuchElementException
except:
    print("selenium library needs to be installed\n")
    os.system("python -m pip install selenium")
    from selenium import webdriver
    from selenium.common.exceptions import NoSuchElementException
try:
    from time import sleep
except:
    print("time library needs to be installed.\n")
    os.system("python -m pip install time")
    
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font
except:
    print("openpyxl library needs to be installed.\n")
    os.system("python -m pip install openpyxl")
    from openpyxl import Workbook
    from openpyxl.styles import Font
    
try:
    import shelve
except:
    print("shelve library needs to be install.\n")
    os.system("python -m pip install shelve")
    import shelve
    
try: 
    from datetime import datetime
except:
    print("datetime library needs to be installed.\n")
    os.system("python -m pip install datetime")
    from datetime import datetime
    
try:
    import sys
except:
    print("sys library needs to be installed.\n")
    os.system("python -m pip install sys")
    import sys

try:
    import pathlib
except:
    print("pathlib library needs to be installed.\n")
    os.system("python -m pip install pathlib")
    import pathlib


FOLDER_LOCATION = pathlib.Path(__file__).parent.absolute()
os.chdir(FOLDER_LOCATION)

num_args = len(sys.argv)
if(num_args == 1): # Normal Execution
    
    ######################################
    ##### Scrape Data from Robinhood #####
    ######################################    

    from helper_functions import make_display, write_detailed_distribution, set_column_width

    opts = webdriver.ChromeOptions()
    opts.add_argument('--log-level=3') # To stop so many messages
    headless = True
    if(headless):
        opts.add_argument('headless')

    browser = webdriver.Chrome('./chromedriver.exe', options = opts)
    browser.get('https://robinhood.com/account')

    loaded = False
    while loaded == False:
        try:
            username_text_box = browser.find_element_by_name('username')
            password_text_box = browser.find_element_by_name('password')
            loaded = True
        except:
            pass
        
    correct_info = False
    while correct_info == False:            
        username = input("Please enter your username.\n")
        username_text_box.send_keys(username)
        
        password = input("Please enter your password.\n")
        password_text_box.send_keys(password)
        
        password_text_box.submit()
        sleep(3)
        try: # Testing if we have successfully left the login page
            browser.find_element_by_name('username')
            username_text_box.clear()
            password_text_box.clear()
        except NoSuchElementException:
            correct_info = True

    leave = False
    while(not leave):
        mfa = input("Enter the multi-factor authentication code.\n")
        mfa_text_box = browser.find_element_by_css_selector(".form-group:nth-child(1) input")
        # mfa_text_box = browser.find_element_by_xpath("//*[@id=\"react_root\"]/div[1]/div[2]/div/div/div[2]/div/div/form/div[2]/div[1]/label/div/input")
        mfa_text_box.send_keys(mfa)
        mfa_text_box.submit()
        sleep(1)
        try:
            browser.find_element_by_css_selector(".form-group:nth-child(1) input")
            mfa_text_box.clear()
        except NoSuchElementException:
            leave = True

    rows = []
    leave = False
    firstPass = False
    while leave == False: # Waiting for page to load
        try:
            # sleep(1)
            rows = browser.find_elements_by_css_selector('.qD5a4psv-CV7GnWdHxLvn._2LZkydeTTkR9XsU9ETIiTB')
            if len(rows) != 0 and firstPass:
                leave = True
            elif len(rows) != 0: # Sometimes it doesn't pickup the cryptocurrencies so I make it double check
                firstPass = True
                sleep(1)
        except:
            pass

    # Using class name of the headers to get each column header
    column_names = browser.find_elements_by_css_selector('._1rWpCWWqvbg316kwNPlwYi .css-17g3u5x')

    # Get all the data
    cells = browser.find_elements_by_css_selector('._97vbM8NcAT1ZQbzihxkmo ._3EYFHtYUcVVUdv_cRQzyNZ , ._97vbM8NcAT1ZQbzihxkmo ._3lBzFefmzZh8nDkfI70QUQ , ._97vbM8NcAT1ZQbzihxkmo ._1htAF8X05ePiHCgFUPbJ3a , ._97vbM8NcAT1ZQbzihxkmo ._2I7disu_qEMm4g6KpT0uvW , ._97vbM8NcAT1ZQbzihxkmo ._21vTQTZz_hYRqoWPrNtlx2 , ._2pdXJ-9MUZn-eEGhEhMPLc , ._1bZB-iudENk38jTXhs7BIB > span , .css-i6t7w2+ span span , ._2jKxrvkjD73sLQEfH5NTgT')

    # Need to read the arrows on the site to tell if "Total Return" is positive or negative
    arrows = browser.find_elements_by_css_selector('.css-i6t7w2 > span > svg > path')
    up_or_down = []
    up_arrow_points = 'M1.5 10L6 2.5L10.5 10L1.5 10Z'
    down_arrow_points = 'M10.5 2L6 9.5L1.5 2L10.5 2Z'
    for i in range(len(arrows)):
        points = arrows[i].get_attribute('d')
        if (points == up_arrow_points):
            up_or_down.append('up')
        elif (points == down_arrow_points):
            up_or_down.append('down')
        else:
            print("ERROR: Unrecognized up or down symbol for stock in row " + str(i))
    

    ##############################
    ##### Create Excel Sheet #####
    ##############################
            
    title_font = Font(size=22, name = 'Cambria', bold = True)
    italic_font = Font(size=14, italic=True)
    smaller_italic_font = Font(size=12, italic=True)
    column_title_font = Font(size = 12, name = 'Cambria')
    nrows = len(rows)
    base_row = 3 # Keep track of the starting row for data entry
    last_row = nrows + base_row - 1

    wb = Workbook()
    sheet = wb.active


    sheet['A1'] = 'Portfolio'
    sheet['J1'] = 'Categories'

    sheet['A1'].font = title_font
    sheet['J1'].font = title_font

    # Adding each column header for Portfolio Data
    n_portfolio_column = 0
    for i in range(1,len(column_names)+1):
        sheet.cell(row = 2, column = i).value = column_names[i-1].text
        sheet.cell(row = 2, column = i).font = column_title_font
        n_portfolio_column += 1


    
    # To use as the index of cells
    counter = 0

    ##### Filling in data for each stock #####
    # Starts at 2 to account for the column names row; Upper Range isn't +3 because rows has 2 extra elements that I want to ignore, these elements exist because I couldn't figure a way to not include them in my search
    for r in range(base_row, last_row + 1): 
        for c in range(1, 8): # There are 7 columns
            if c == 6: # If it is on the 'Total Return' column (which can be negative or positive but robinhood uses an arrow to display that)
                if up_or_down[r-3] == 'down':
                    sheet.cell(row = r, column = c).value = float('-' + (cells[counter].text[1:]).replace(',', ''))   
                else:
                    sheet.cell(row = r, column = c).value = float((cells[counter].text[1:]).replace(',', ''))                
            elif c < 3:
                sheet.cell(row = r, column = c).value = cells[counter].text
            elif c == 3:
                sheet.cell(row = r, column = c).value = float((cells[counter].text).replace(',', ''))
            else:
                sheet.cell(row = r, column = c).value = float((cells[counter].text[1:]).replace(',', ''))
            counter += 1

    browser.close() # Don't need the browser anymore    

    # Adding a cell to display current worth
    sheet.cell(row = last_row + 1, column = 7).value = '=ROUND(SUM(G'+str(base_row)+':G'+str(last_row)+'), 2)'
    sheet.cell(row = last_row + 1, column = 7).font = italic_font

    # Adding a cell to display gains/loses 
    sheet.cell(row = last_row + 1, column = 6).value = '=SUM(F'+str(base_row)+':F'+str(last_row)+')'
    sheet.cell(row = last_row + 1, column = 6).font = italic_font

    # Adding a cell to display initial investment
    sheet.cell(row = last_row + 1, column = 5).value = '=G'+str(last_row + 1)+'-F'+str(last_row + 1)
    sheet.cell(row = last_row + 1, column = 5).font = italic_font

    # Calculating total equity, total gain/lose, and initial investment
    total_return = 0
    for i in range(base_row, last_row + 1):
        total_return += sheet.cell(row =  i, column = 6).value
    equity = 0
    for i in range(base_row, last_row + 1):
        equity += sheet.cell(row =  i, column = 7).value
    initial_investment = 0
    for i in range(base_row, last_row + 1):
        initial_investment += sheet.cell(row =  i, column = 5).value * sheet.cell(row = i, column = 3).value

    # Creating a column for Percent Change
    sheet.cell(row = base_row - 1, column = 8).value = '% Change'
    for r in range(base_row, last_row + 1):
        try:
            # Total return / (av cost * shares)
            sheet.cell(row = r, column = 8).value = round(sheet.cell(row = r, column = 6).value/(sheet.cell(row = r, column = 5).value * sheet.cell(row = r, column = 3).value)*100, 2)
        except:
            sheet.cell(row = r, column = 8).value = 'Error Division by Zero'
            pass
    
    sheet.cell(row = last_row + 1, column = 8).value = round((total_return/initial_investment)*100, 2)
    sheet.cell(row = last_row + 1, column = 8).font = italic_font
    n_portfolio_column += 1
    
    # Database to keep track of which stocks are small cap, mid cap, etc...
    designations = shelve.open('Stock_Categories')
    n_small = []
    n_mid = []
    n_large = []
    n_all = []
    n_value = []
    n_growth = []
    n_mixed = []
    n_inter = []
    n_emerging = []
    n_crypto = []

    ##### Adding each stock I have to their respective list #####
    for row in range(base_row, last_row + 1):
        ticker = sheet.cell(row = row, column = 2).value
        try:
            if designations[ticker][0] == 'L':
                n_large.append(ticker)
            elif designations[ticker][0] == 'M':
                n_mid.append(ticker)
            elif designations[ticker][0] == 'S':
                n_small.append(ticker)
            elif designations[ticker][0] == 'A':
                n_all.append(ticker)
            elif designations[ticker][0] == 'E':
                n_emerging.append(ticker)

            if designations[ticker][1] == 'G':
                n_growth.append(ticker)
            elif designations[ticker][1] == 'V':
                n_value.append(ticker)
            elif designations[ticker][1] == 'M':
                n_mixed.append(ticker)
            elif designations[ticker][1] == 'I':
                n_inter.append(ticker)
            
            if designations[ticker][2] == "CRYP":
                n_crypto.append(ticker)
        except:
            print(ticker + ' has no data in database.')
            pass

    #######################################################
    ##### Display performance by each group of stocks #####
    #######################################################
    
    n_display_column = 0
    columns_per_display = 5 # This needs to be updated if I add a column to display

    column = n_portfolio_column + 2 # Column value for first column of display
    make_display(2, column, n_large, 'Large Cap', sheet, nrows, base_row, last_row)
    make_display(len(n_large) + 5, column, n_mid, 'Mid Cap', sheet, nrows, base_row, last_row)
    make_display(len(n_large) + len(n_mid) + 8, column, n_small, 'Small Cap', sheet, nrows, base_row, last_row)
    make_display(len(n_large) + len(n_mid) + len(n_small) + 11, column, n_inter, 'Inter.', sheet, nrows, base_row, last_row)
    make_display(len(n_large) + len(n_mid) + len(n_small) + len(n_inter) + 14, column, n_crypto, "Crypto", sheet, nrows, base_row, last_row)
    n_display_column += columns_per_display

    column = n_portfolio_column + n_display_column + 3 # Column value for second column of display
    make_display(2, column, n_value, 'Value', sheet, nrows, base_row, last_row)
    make_display(len(n_value) + 5, column, n_mixed, 'Mixed', sheet, nrows, base_row, last_row)
    make_display(len(n_value) + len(n_mixed) + 8, column, n_growth, 'Growth', sheet, nrows, base_row, last_row)
    make_display(len(n_value) + len(n_mixed) + len(n_growth) + 11, column, n_inter, 'Inter.', sheet, nrows, base_row, last_row)
    n_display_column += columns_per_display

    ###############################################
    ##### Display amount in each category and #####
    ##### how much to invest in each category #####
    ###############################################

    write_detailed_distribution(n_portfolio_column + n_display_column + 4, sheet, designations, nrows, base_row, last_row)

    designations.close()
    
    ##########################
    ##### Resize Columns #####
    ##########################
    set_column_width('A',   35, sheet)
    set_column_width('B',   7,  sheet)
    set_column_width('C',   7,  sheet)
    set_column_width('D',   7,  sheet)
    set_column_width('E',   13, sheet)
    set_column_width('F',   13, sheet)
    set_column_width('G',   13, sheet)
    set_column_width('H',   8., sheet)
    set_column_width('I',   2,  sheet)
    set_column_width('J',   10, sheet)
    set_column_width('K',   7,  sheet)
    set_column_width('L',   13, sheet)
    set_column_width('M',   8,  sheet)
    set_column_width('N',   9,  sheet)
    set_column_width('O',   2,  sheet)
    set_column_width('P',   10, sheet)
    set_column_width('Q',   7,  sheet)
    set_column_width('R',   12, sheet)
    set_column_width('S',   8,  sheet)
    set_column_width('T',   9,  sheet)
    set_column_width('U',   2,  sheet)
    set_column_width('V',   18, sheet)
    set_column_width('W',   10, sheet)
    set_column_width('X',   6,  sheet)
    set_column_width('Y',   6,  sheet)
    set_column_width('Z',   10, sheet)
    set_column_width('AA',  8,  sheet)
    set_column_width('AB',  10, sheet)
    set_column_width('AC',  7,  sheet)
    set_column_width('AD',  12, sheet)

    ############################
    ##### Ending procedure #####
    ############################

    # Time stamp
    sheet.cell(row = last_row + 2, column = 1).value = 'Date and time created:'
    sheet.cell(row = last_row + 3, column = 1).value = datetime.now()
    
    # Creating a variable to be used to name output file
    FILE_NAME = username + "_portfolio.xlsx"

    print("Trying to save to:\t" + FILE_NAME)
    saved = False
    while not saved: # Allow them to close the file if it is already open
        try:        
            wb.save(FILE_NAME)
            saved = True
        except PermissionError:
            print("Permission denied\nTry closing the file to allow it to be saved.\n")
            input("Press any key when you have closed the file.\n")
    print("Successfully saved to:\t" + FILE_NAME)    
    
    wb.close()
    
elif num_args == 2 and sys.argv[1] == 'addStock':
    # If you want to add a stock to the database you need to specify the type
    print('adding stock')
    ticker      = input('What is the ticker of the stock?\n')
    stockCat    = input(
    """Describe the stock: (Enter the value in the parenthesis.)
    United States:      (TSM)
    Large Blend:        (LCB)
    Large Value:        (LCV)
    Large Growth:       (LCG)
    Mid Blend:          (MCB)
    Mid Value:          (MCV)
    Mid Growth:         (MCG)
    Small Blend:        (SCB)
    Small Value:        (SCV)
    Small Growth:       (SCG)
    Total Inter.:       (TSMI)
    ex-US Large Blend:  (LCBI)
    ex-US Large Value:  (LCVI)
    ex-US Large Growth: (LCGI)
    ex-US Mid Blend:    (MCBI)
    ex-US Mid Value:    (MCVI)
    ex-US Mid Growth:   (MCGI)
    ex-US Small Blend:  (SCBI)
    ex-US Small Value:  (SCVI)
    ex-US Small Growth: (SCGI)
    Emerging:           (EM)
    Long Term Bonds:    (LTB)
    Intermediate Bonds: (ITB)
    Short Term Bonds:   (STB)
    Treasury Bills:     (TB)
    Commodities         (COM)
    REITs               (REIT)
    Gold                (GLD)
    Cryptocurrency      (CRYP)
    """)
    if(stockCat == 'TSM'): 
        size    = 'L'
        quality = 'M'
    elif(stockCat == 'LCB'):
        size    = 'L'
        quality = 'M'
    elif(stockCat == 'LCV'):
        size    = 'L'
        quality = 'V'
    elif(stockCat == 'LCG'):
        size    = 'L'
        quality = 'G'
    elif(stockCat == 'MCB'):
        size    = 'M'
        quality = 'M'
    elif(stockCat == 'MCV'):
        size    = 'M'
        quality = 'V'
    elif(stockCat == 'MCG'):
        size    = 'M'
        quality = 'G'
    elif(stockCat == 'SCB'):
        size    = 'S'
        quality = 'M'
    elif(stockCat == 'SCV'):
        size    = 'S'
        quality = 'V'
    elif(stockCat == 'SCG'):
        size    = 'S'
        quality = 'G'
    elif(stockCat == 'TSMI'):
        size    = 'A'
        quality = 'I'
    elif(stockCat == 'LCBI'):
        size    = 'L'
        quality = 'I'
    elif(stockCat == 'LCVI'):
        size    = 'L'
        quality = 'I'
    elif(stockCat == 'LCGI'):
        size    = 'L'
        quality = 'I'
    elif(stockCat == 'MCBI'):
        size    = 'M'
        quality = 'I'
    elif(stockCat == 'MCVI'):
        size    = 'M'
        quality = 'I'
    elif(stockCat == 'MCGI'):
        size    = 'M'
        quality = 'I'
    elif(stockCat == 'SCBI'):
        size    = 'S'
        quality = 'I'
    elif(stockCat == 'SCVI'):
        size    = 'S'
        quality = 'I'
    elif(stockCat == 'SCGI'):
        size    = 'S'
        quality = 'I'
    elif(stockCat == 'EM'):
        size    = 'E'
        quality = 'I'
    elif(stockCat in ('LTB', 'ITB', 'STB', 'TB', 'COM', 'REIT', 'GLD', 'CRYP')):
        size    = 'NA'
        quality = 'NA'
    else:
        print('It seems that you have entered information that we are unable to process.')
        sys.exit()
    
    designations = shelve.open(os.path.join(FOLDER_LOCATION, 'Stock_Categories'))
    print('\nYou are about to add this to the database:\nTicker: ' + ticker + '\nSize: ' + size + '\nQuality: ' + quality + '\nCategory: ' + stockCat + '\n')
    cont = input('Is this Ok?(y/n)\n')
    if cont != 'y':
        print('Editing Cancled')
        sys.exit()
    designations[ticker] = [size, quality, stockCat]
    print('Added')
    pass
else:
    print('Please note that this program currently only supports one command line argument: \"addStock\"')
